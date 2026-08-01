import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

ATTENDANCE_FOLDER = "Attendance"


def get_today_file():
    """Returns today's attendance file path."""
    os.makedirs(ATTENDANCE_FOLDER, exist_ok=True)

    today = datetime.now().strftime("%Y-%m-%d")
    return os.path.join(ATTENDANCE_FOLDER, f"Attendance_{today}.xlsx")


def create_file_if_not_exists():
    """Creates today's attendance file if missing."""

    file = get_today_file()

    if not os.path.exists(file):

        wb = Workbook()

        ws = wb.active

        ws.title = "Attendance"

        ws.append([
            "Person ID",
            "Name",
            "Date",
            "Time",
            "Status"
        ])

        wb.save(file)

    return file


def already_marked(person_id):
    """Returns True if attendance already marked today."""

    file = create_file_if_not_exists()

    wb = load_workbook(file)

    ws = wb.active

    for row in ws.iter_rows(min_row=2):

        if row[0].value == person_id:
            return True

    return False


def mark_attendance(person_id, name):

    file = create_file_if_not_exists()

    wb = load_workbook(file)

    ws = wb.active

    now = datetime.now()

    ws.append([
        person_id,
        name,
        now.strftime("%Y-%m-%d"),
        now.strftime("%H:%M:%S"),
        "Present"
    ])

    wb.save(file)