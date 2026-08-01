import os
from datetime import datetime
from openpyxl import load_workbook

ATTENDANCE_FOLDER = "Attendance"


def get_today_file():
    today = datetime.now().strftime("%Y-%m-%d")
    return os.path.join(ATTENDANCE_FOLDER, f"Attendance_{today}.xlsx")


while True:

    print("=" * 60)
    print("ATTENDANCE REPORT")
    print("=" * 60)

    print("1. Today's Attendance")
    print("2. Search by Date")
    print("3. Show All")
    print("4. Back")

    choice = input("\nEnter Choice : ")

    file = get_today_file()

    if not os.path.exists(file):
        print("\nToday's attendance file not found.")
        input("\nPress Enter...")
        break

    wb = load_workbook(file)

    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))

    if len(rows) <= 1:
        print("\nNo attendance records found.")
        input("\nPress Enter...")
        break

    if choice == "1":

        today = datetime.now().strftime("%Y-%m-%d")

        print(f"\n{'ID':<10}{'Name':<25}{'Date':<15}{'Time':<12}{'Status'}")
        print("-"*75)

        count = 0

        for row in rows[1:]:

            if row[2] == today:

                print(f"{row[0]:<10}{row[1]:<25}{row[2]:<15}{row[3]:<12}{row[4]}")

                count += 1

        print("-"*75)
        print("Records :", count)

    elif choice == "2":

        search = input("Enter Date (YYYY-MM-DD): ")

        print(f"\n{'ID':<10}{'Name':<25}{'Date':<15}{'Time':<12}{'Status'}")
        print("-"*75)

        count = 0

        for row in rows[1:]:

            if row[2] == search:

                print(f"{row[0]:<10}{row[1]:<25}{row[2]:<15}{row[3]:<12}{row[4]}")

                count += 1

        print("-"*75)
        print("Records :", count)

    elif choice == "3":

        print(f"\n{'ID':<10}{'Name':<25}{'Date':<15}{'Time':<12}{'Status'}")
        print("-"*75)

        for row in rows[1:]:

            print(f"{row[0]:<10}{row[1]:<25}{row[2]:<15}{row[3]:<12}{row[4]}")

        print("-"*75)
        print("Total Records :", len(rows)-1)

    elif choice == "4":
        break

    else:
        print("Invalid Choice")

    input("\nPress Enter...")